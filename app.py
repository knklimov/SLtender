
import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill
import io

# Цвета в формате ARGB
TARGET_GREEN = "FF5AFC4C"
TARGET_YELLOW = "FFFCF74C"
TARGET_RED = "FFFC4C4C"

def process_excel(data_file, plan_file, yellow_step, red_step):
    wb_data = openpyxl.load_workbook(data_file)
    wb_plan = openpyxl.load_workbook(plan_file)
    modifications = []

    for sheet_name in wb_data.sheetnames:
        if sheet_name not in wb_plan.sheetnames:
            st.warning(f"Лист '{sheet_name}' отсутствует в файле с минимумами. Пропускаем.")
            continue

        ws_data = wb_data[sheet_name]
        ws_plan = wb_plan[sheet_name]

        if ws_data.max_row != ws_plan.max_row or ws_data.max_column != ws_plan.max_column:
            st.warning(f"Размеры листа '{sheet_name}' не совпадают. Пропускаем.")
            continue

        for row in ws_data.iter_rows(min_row=2):
            for cell in row:
                fill = cell.fill
                if fill and fill.fill_type == "solid" and fill.fgColor and fill.fgColor.rgb:
                    color = fill.fgColor.rgb
                    if not (color and isinstance(color, str)):
                        continue
                    color = color.upper()
                    if color == TARGET_GREEN:
                        continue

                    try:
                        original_value = cell.value
                        min_cell = ws_plan[cell.coordinate]
                        min_value = min_cell.value

                        if not (isinstance(original_value, (int, float)) and isinstance(min_value, (int, float))):
                            continue

                        if color == TARGET_YELLOW:
                            new_value = max(original_value - yellow_step, min_value)
                        elif color == TARGET_RED:
                            new_value = max(original_value - red_step, min_value)
                        else:
                            continue  # не известный цвет

                        if new_value != original_value:
                            cell.value = new_value
                            modifications.append((sheet_name, cell.coordinate, original_value, new_value, color, min_value))

                    except Exception:
                        pass

    output = io.BytesIO()
    wb_data.save(output)
    output.seek(0)
    return output, modifications

st.title("Редактирование Excel по цвету с разными шагами и минимумами")

yellow_step = st.number_input("Шаг снижения для ЖЁЛТОГО цвета", min_value=1, max_value=10000, value=3)
red_step = st.number_input("Шаг снижения для КРАСНОГО цвета", min_value=1, max_value=10000, value=5)

plan_file = st.file_uploader("🔢 Загрузите файл с минимальными значениями", type=["xlsx"])
data_file = st.file_uploader("🎯 Загрузите файл для редактирования (цвета)", type=["xlsx"])



if plan_file and data_file:
    processed_file, log = process_excel(data_file, plan_file, yellow_step, red_step)

    st.download_button(
        label="📥 Скачать обработанный файл",
        data=processed_file,
        file_name="output_modified.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    if log:
        st.write("Изменённые ячейки:")
        st.dataframe(
            {"Лист": [x[0] for x in log],
             "Ячейка": [x[1] for x in log],
             "Старое значение": [x[2] for x in log],
             "Новое значение": [x[3] for x in log],
             "Цвет": [x[4] for x in log],
             "Минимум": [x[5] for x in log]}
        )
    else:
        st.success("Файл обработан: изменений не найдено.")
